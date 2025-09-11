import base64, io, os, qrcode, urllib
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Ranking, Torneio


@login_required(redirect_field_name='next', login_url='/admin/login/')
def create_games(request, torneio_id: str):
    '''Cria jogos para o torneio'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    result = torneio.create_games()
    if isinstance(result, Exception):
        messages.add_message(request, messages.ERROR, result)
    else:
        messages.add_message(request, messages.INFO, f'{ len(result )} jogos gerados com sucesso!')
    response = redirect('admin:bt_league_torneio_change', torneio_id)
    response['location'] += '#jogos-tab'
    return response


def finish_tournament(request, torneio_id: str):
    '''Finaliza torneio e desabilita todas os jogadores'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    torneio.finish()
    messages.add_message(request, messages.SUCCESS, f'{ torneio } finalizado!')
    response = redirect('admin:bt_league_torneio_change', torneio_id)
    response['location'] += '#jogos-tab'
    return response


def see_tournament(request, torneio_id: str):
    '''Visualiza torneio'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    jogos = torneio.jogo_set.all()

    # Calcular ranking
    ranking = []
    for jogador in torneio.jogadores.all():
        vitorias, pontos, saldo, n_jogos = jogador.player_points(torneio)
        ranking.append({
            'jogador': jogador,
            'pontos': pontos,
            'saldo': saldo,
            'vitorias': vitorias,
            'jogos': n_jogos
        })

    # Ordenar ranking por posição
    ranking = sorted(ranking, key=lambda x: (-x['vitorias'], -x['saldo'], -x['pontos']))
    ranking_result = []
    j_0 = {'pontos': 0, 'saldo': 0, 'vitorias': 0, 'posicao': 1}
    for i, j_1 in enumerate(ranking):
        if (j_1['pontos'] == j_0['pontos']) and (j_1['saldo'] == j_0['saldo']) and (j_1['vitorias'] == j_0['vitorias']):
            j_1['posicao'] = j_0['posicao']
        else:
            j_1['posicao'] = i + 1
            j_0 = j_1
        ranking_result.append(j_1)

    context = {
        'torneio': torneio,
        'jogos': jogos,
        'n_jogos': jogos.count(),
        'jogos_restantes': jogos.exclude(concluido='C').count(),
        'ranking': ranking_result,
        'can_edit': request.user.is_superuser or torneio.criado_por == request.user,
    }
    return render(request, 'bt_league/see_league.html', context)


@login_required(redirect_field_name='next', login_url='/admin/login/')
def qrcode_tournament(request, torneio_id: str):
    '''Cria QR Code do torneio'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    img = qrcode.make(f'{ os.getenv("APP_LINK") }/rei-rainha/{ torneio.slug }')
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    string = base64.b64encode(buf.read())
    uri =  urllib.parse.quote(string)
    context = {
        'img_b64': uri,
        'torneio': torneio,
    }
    return render(request, 'qrcode_tournament.html', context)


def export_csv(request, torneio_id: str):
    '''Gera um arquivo CSV com os dados dos jogadores'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    torneio_jogos = torneio.jogo_set.all()

    wb = Workbook()
    ws = wb.active
    ws.title = torneio.nome

    # Estilos de formatação
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    normal_font = Font(size=10)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    orange_fill = PatternFill(start_color='FF972F', end_color='FF972F', fill_type='solid')
    game_headers = ['', 'Dupla 1', 'Placar', '', '', 'Dupla 2']
    ranking_headers = ['#', 'Jogador', 'Vitórias', 'Saldo', 'Pontos', 'Jogos']
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # LINHA 1: Nome do torneio
    ws.merge_cells('A1:N1')
    ws['A1'] = torneio.nome
    ws['A1'].font = title_font
    ws['A1'].fill = orange_fill
    ws['A1'].alignment = center_alignment
    ws['A1'].border = thin_border
    ws['N1'].border = thin_border
    rd = ws.row_dimensions[1]
    rd.height = 45

    image_path = settings.BASE_DIR / 'setup/static/images/trophy.png'
    if os.path.exists(image_path):
        try:
            img = Image(image_path)
            img.width = 60
            img.height = 60
            ws.add_image(img, 'A1')
        except Exception as e:
            print(f"Erro ao carregar imagem: {e}")

    # LINHA 2: Data do torneio
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Data: {torneio.data.strftime('%d/%m/%Y')}"
    ws['A2'].font = header_font
    ws['A2'].fill = orange_fill
    ws['A2'].alignment = center_alignment
    ws['A2'].border = thin_border
    ws['F2'].border = thin_border

    ws.merge_cells('G2:H2')
    ws['G2'].fill = orange_fill

    ws.merge_cells('I2:N2')
    ws['I2'] = f"Jogos: {torneio_jogos.count()}"
    ws['I2'].font = header_font
    ws['I2'].fill = orange_fill
    ws['I2'].alignment = center_alignment
    ws['I2'].border = thin_border
    ws['N2'].border = thin_border

    # LINHA 6: Headers da tabela
    for col, header in enumerate(ranking_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Preparar dados do ranking
    ranking = []
    for jogador in torneio.jogadores.all():
        vitorias, pontos, saldo, jogos = jogador.player_points(torneio)
        ranking.append({
            'jogador': jogador,
            'vitorias': vitorias,
            'pontos': pontos,
            'saldo': saldo,
            'jogos': jogos
        })

    # Ordenar ranking por posição
    ranking = sorted(ranking, key=lambda x: (-x['vitorias'], -x['saldo'], -x['pontos']))

    # Calcular posições (considerando empates)
    ranking_result = []
    j_0 = {'pontos': 0, 'saldo': 0, 'vitorias': 0, 'posicao': 1}
    for i, j_1 in enumerate(ranking):
        if (j_1['pontos'] == j_0['pontos']) and (j_1['saldo'] == j_0['saldo']) and (j_1['vitorias'] == j_0['vitorias']):
            j_1['posicao'] = j_0['posicao']
        else:
            j_1['posicao'] = i + 1
            j_0 = j_1
        ranking_result.append(j_1)

    # LINHA 7+: Adicionar dados dos jogadores
    for row_idx, jogador in enumerate(ranking_result, 6):
        ws.cell(row=row_idx, column=1, value=jogador['posicao'])
        ws.cell(row=row_idx, column=2, value=jogador['jogador'].nome)
        ws.cell(row=row_idx, column=3, value=jogador['vitorias'])
        ws.cell(row=row_idx, column=4, value=jogador['saldo'])
        ws.cell(row=row_idx, column=5, value=jogador['pontos'])
        ws.cell(row=row_idx, column=6, value=jogador['jogos'])

        for col in [1, 2, 3, 4, 5, 6]:
            ws.cell(row=row_idx, column=col).alignment = center_alignment
            ws.cell(row=row_idx, column=col).border = thin_border
            ws.cell(row=row_idx, column=col).font = normal_font

    for col, header in enumerate(game_headers, 9):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    rd = ws.row_dimensions[6]
    rd.height = 30

    ws.merge_cells(
        start_column=11,
        start_row=5,
        end_column=13,
        end_row=5
    )

    for row_idx, jogo in enumerate(torneio_jogos, 6):
        ws.cell(row=row_idx, column=9, value=jogo.get_concluido_display())
        ws.cell(row=row_idx, column=10, value=jogo.render_dupla_1())
        ws.cell(row=row_idx, column=11, value=jogo.placar_dupla1)
        ws.cell(row=row_idx, column=12, value='x')
        ws.cell(row=row_idx, column=13, value=jogo.placar_dupla2)
        ws.cell(row=row_idx, column=14, value=jogo.render_dupla_2())

        for col in [9, 10, 11, 12, 13, 14]:
            ws.cell(row=row_idx, column=col).alignment = center_alignment
            ws.cell(row=row_idx, column=col).border = thin_border
            ws.cell(row=row_idx, column=col).font = normal_font

        rd = ws.row_dimensions[row_idx]
        rd.height = 30

    column_widths = {
        'A': 3,
        'B': 15,
        'C': 8,
        'D': 8,
        'E': 8,
        'F': 8,
        'G': 5,
        'H': 5,
        'I': 5,
        'J': 15,
        'K': 3,
        'L': 3,
        'M': 3,
        'N': 15
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # SALVAR EM MEMÓRIA
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # CRIAR RESPONSE
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{torneio.nome}.xlsx"'

    return response


def see_ranking(request, ranking_id: str):
    '''Visualiza o ranking geral dos jogadores em torneios'''
    ranking = get_object_or_404(Ranking, pk=ranking_id)
    torneios = Torneio.objects.filter(ranking=ranking)
    jogadores = {}
    for torneio in torneios:
        for jogador in torneio.jogadores.all():
            vitorias, pontos, saldo, jogos = jogador.player_points(torneio)
            if jogador.id not in jogadores:
                jogadores[jogador.id] = {
                    'jogador': jogador.nome,
                    'vitorias': vitorias,
                    'saldo': saldo,
                    'pontos': pontos,
                    'jogos': jogos
                }
            else:
                jogadores[jogador.id]['vitorias'] += vitorias
                jogadores[jogador.id]['saldo'] += saldo
                jogadores[jogador.id]['pontos'] += pontos
                jogadores[jogador.id]['jogos'] += jogos

    jogadores = list(jogadores.values())
    jogadores = sorted(jogadores, key=lambda x: (-x['vitorias'], -x['saldo'], -x['pontos'], x['jogos']))
    jogadores_ranking = []
    j_0 = {'pontos': 0, 'saldo': 0, 'vitorias': 0, 'posicao': 1}
    for i, j_1 in enumerate(jogadores):
        if (j_1['pontos'] == j_0['pontos']) and (j_1['saldo'] == j_0['saldo']) and (j_1['vitorias'] == j_0['vitorias']):
            j_1['posicao'] = j_0['posicao']
        else:
            j_1['posicao'] = i + 1
            j_0 = j_1
        jogadores_ranking.append(j_1)

    context = {
        'ranking': ranking,
        'jogadores': jogadores_ranking
    }
    return render(request, 'bt_league/see_ranking.html', context)


def get_tournament_data(request, torneio_id: str):
    '''Returns tournament data as JSON for React frontend'''
    torneio = Torneio.objects.filter(slug=torneio_id).first()
    if not torneio:
        return JsonResponse({'error': 'Torneio não encontrado'}, status=404)
    jogos = torneio.jogo_set.all()

    # Calculate ranking
    ranking = []
    for jogador in torneio.jogadores.all():
        vitorias, pontos, saldo, n_jogos = jogador.player_points(torneio)
        ranking.append({
            'jogador': jogador.nome,
            'pontos': pontos,
            'saldo': saldo,
            'vitorias': vitorias,
            'jogos': n_jogos
        })

    # Sort ranking by position
    ranking = sorted(ranking, key=lambda x: (-x['vitorias'], -x['saldo'], -x['pontos']))
    ranking_result = []
    j_0 = {'pontos': 0, 'saldo': 0, 'vitorias': 0, 'posicao': 1}
    for i, j_1 in enumerate(ranking):
        if (j_1['pontos'] == j_0['pontos']) and (j_1['saldo'] == j_0['saldo']) and (j_1['vitorias'] == j_0['vitorias']):
            j_1['posicao'] = j_0['posicao']
        else:
            j_1['posicao'] = i + 1
            j_0 = j_1
        ranking_result.append(j_1)

    data = {
        'torneio': {
            'id': torneio.id,
            'nome': torneio.nome,
            'data': torneio.data,
            'ativo': torneio.ativo,
        },
        'jogos': [
            {
                'id': jogo.id,
                'dupla1': jogo.dupla_1(),
                'dupla2': jogo.dupla_2(),
                'placar_dupla1': jogo.placar_dupla1,
                'placar_dupla2': jogo.placar_dupla2,
                'concluido': jogo.concluido,
                'quadra': jogo.quadra
            } for jogo in jogos
        ],
        'estatisticas': {
            'total_jogos': jogos.count(),
            'jogos_restantes': jogos.exclude(concluido='C').count(),
        },
        'ranking': ranking_result,
        'can_edit': request.user.is_superuser or torneio.criado_por == request.user,
    }

    return JsonResponse(data)