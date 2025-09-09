import base64, io, os, qrcode, urllib
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Jogo, Torneio


def distribute_classifieds(classifieds):
    '''Distribui classificados [1,2,3,4,5,6,7,8,9,10] -> [1,10,2,9,3,8,4,7,5,6]'''
    result = []
    for i in range(0, len(classifieds), 2):
        result.extend([classifieds[i], classifieds[-i-1]])
    return result


@login_required(redirect_field_name='next', login_url='/admin/login/')
def create_games(request, torneio_id: str):
    '''Cria jogos para o torneio'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    result = torneio.create_games()
    if isinstance(result, Exception):
        messages.add_message(request, messages.ERROR, result)
    else:
        messages.add_message(request, messages.INFO, 'Jogos gerados com sucesso!')
    response = redirect('admin:bt_cup_torneio_change', torneio_id)
    response['location'] += '#jogos-tab'
    return response


def finish_tournament(request, torneio_id: str):
    '''Finaliza torneio e desabilita todas as duplas'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    torneio.finish()
    messages.add_message(request, messages.SUCCESS, f'{ torneio } finalizado!')
    response = redirect('admin:bt_cup_torneio_change', torneio_id)
    response['location'] += '#jogos-tab'
    return response


@login_required(redirect_field_name='next', login_url='/admin/login/')
def next_stage(request, torneio_id: str):
    '''Processa grupos e vai para a próxima fase'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    jogos = torneio.jogo_set.all()
    jogos_preenchidos = jogos.filter(
        fase__in=['OITAVAS', 'QUARTAS', 'SEMIFINAIS', 'FINAL'],
        dupla1__isnull=False,
        dupla2__isnull=False
    )
    jogos_grupos_nao_finalizados = jogos.filter(fase__startswith='GRUPO').exclude(concluido='C')

    if jogos_grupos_nao_finalizados.exists():
        messages.add_message(request, messages.ERROR, 'Jogos de grupos ainda não foram finalizados.')
    elif jogos_preenchidos:
        result = torneio.next_stage()
        if isinstance(result, Exception):
            messages.add_message(request, messages.ERROR, result)
        else:
            messages.add_message(request, messages.INFO, 'Confrontos gerados com sucesso!')
    else:
        classificados = torneio.process_groups()
        classificados = distribute_classifieds(classificados)
        for i in range(0, len(classificados), 2):
            dupla1, dupla2 = classificados[i:i+2]
            if torneio.quantidade_grupos == 1:
                final = jogos.get(fase='FINAL')
                final.dupla1 = dupla1['dupla']
                final.dupla2 = dupla2['dupla']
                final.save()
            elif torneio.quantidade_grupos == 2:
                semifinal = jogos.filter(fase='SEMIFINAIS')[i//2]
                semifinal.dupla1 = dupla1['dupla']
                semifinal.dupla2 = dupla2['dupla']
                semifinal.save()
            elif torneio.quantidade_grupos == 4:
                quartas = jogos.filter(fase='QUARTAS')[i//2]
                quartas.dupla1 = dupla1['dupla']
                quartas.dupla2 = dupla2['dupla']
                quartas.save()
            elif torneio.quantidade_grupos == 8:
                oitavas = jogos.filter(fase='OITAVAS')[i//2]
                oitavas.dupla1 = dupla1['dupla']
                oitavas.dupla2 = dupla2['dupla']
                oitavas.save()
        messages.add_message(request, messages.INFO, 'Confrontos gerados com sucesso!')

    response = redirect('admin:bt_cup_torneio_change', torneio_id)
    response['location'] += '#jogos-tab'
    return response


def see_tournament(request, torneio_id: str):
    '''Visualiza torneio'''
    CARD_STYLE_DICT = {
        'OITAVAS': 'col-sm-3',
        'QUARTAS': 'col-sm-4',
        'SEMIFINAIS': 'col-sm-6',
        'FINAL': 'col-sm-6',
    }
    torneio = Torneio.objects.get(id=torneio_id)
    jogos = Jogo.objects.filter(torneio=torneio)
    classificacao = torneio.get_groups_ranking()
    playoff_card_style = ''

    # Separar jogos por grupos
    grupos = {}
    groups_finished = True
    for i in range(1, 5):  # Grupos 1 a 4
        grupo_jogos = jogos.filter(fase=f'GRUPO {i}')
        if grupo_jogos.exists():
            grupos[f'GRUPO {i}'] = {
                'jogos': grupo_jogos,
                'classificacao': classificacao[f'GRUPO {i}']
            }
            not_finished = grupo_jogos.exclude(concluido='C').exists()
            groups_finished = groups_finished and not not_finished

    # Separar fases finais (semifinais e final)
    fases_finais = {}
    for fase in ['FINAL', 'SEMIFINAIS', 'QUARTAS', 'OITAVAS']:
        fase_jogos = jogos.filter(fase=fase)
        if fase_jogos.exists():
            fases_finais[fase] = fase_jogos
            playoff_card_style = CARD_STYLE_DICT[fase]

    context = {
        'torneio': torneio,
        'grupos': grupos,
        'fases_finais': fases_finais,
        'classificacao': classificacao,
        'groups_finished': groups_finished,
        'playoff_card_style': playoff_card_style,
        'can_edit': request.user.is_superuser or torneio.criado_por == request.user,
    }
    return render(request, 'bt_cup/see_cup.html', context)


@login_required(redirect_field_name='next', login_url='/admin/login/')
def qrcode_tournament(request, torneio_id: str):
    '''Cria QR Code do torneio'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    site_url = os.getenv('HOST_ADDRESS')
    # img = qrcode.make(f'{ site_url }/torneio/{ torneio_id }')
    img = qrcode.make(f'{ os.getenv("APP_LINK") }/torneio/{ torneio_id }')
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    string = base64.b64encode(buf.read())
    uri =  urllib.parse.quote(string)
    context = {
        'img_b64': uri,
        'torneio': torneio,
    }
    return render(request, 'qrcode_tournament.html', context)


def get_tournament_data(request, torneio_id: str):
    '''Returns tournament data as JSON for React frontend'''
    CARD_STYLE_DICT = {
        'OITAVAS': '1/4',
        'QUARTAS': '1/3',
        'SEMIFINAIS': '1/2',
        'TERCEIRO LUGAR': '1/2',
        'FINAL': '1/2',
    }
    torneio = Torneio.objects.filter(pk=torneio_id).first()
    if not torneio:
        return JsonResponse({'error': 'Torneio não encontrado'}, status=404)
    jogos = Jogo.objects.filter(torneio=torneio)
    classificacao = torneio.get_groups_ranking()
    playoff_card_style = ''

    # Process groups
    grupos = {}
    groups_finished = True
    for i in range(1, 5):
        grupo_jogos = jogos.filter(fase=f'GRUPO {i}')
        if grupo_jogos.exists():
            grupos[f'GRUPO {i}'] = {
                'jogos': [
                    {
                        'id': jogo.id,
                        'dupla1': jogo.dupla1.render() if jogo.dupla1 else None,
                        'dupla2': jogo.dupla2.render() if jogo.dupla2 else None,
                        'placar_dupla1': jogo.placar_dupla1,
                        'placar_dupla2': jogo.placar_dupla2,
                        'concluido': jogo.concluido,
                    } for jogo in grupo_jogos
                ],
                'classificacao': [
                    {
                        'posicao': dupla['posicao'],
                        'dupla': dupla['dupla'].render(),
                        'vitorias': dupla['vitorias'],
                        'pontos': dupla['pontos'],
                        'saldo': dupla['saldo'],
                        'jogos': dupla['jogos'],
                    } for dupla in classificacao[f'GRUPO {i}']
                ]
            }
            groups_finished = groups_finished and not grupo_jogos.exclude(concluido='C').exists()

    # Process playoffs
    fases_finais = {}
    for fase in ['FINAL', 'TERCEIRO LUGAR', 'SEMIFINAIS', 'QUARTAS', 'OITAVAS']:
        fase_jogos = jogos.filter(fase=fase)
        if fase_jogos.exists():
            fases_finais[fase] = [
                {
                    'id': jogo.id,
                    'dupla1': jogo.dupla1.render() if jogo.dupla1 else None,
                    'dupla2': jogo.dupla2.render() if jogo.dupla2 else None,
                    'placar_dupla1': jogo.placar_dupla1,
                    'placar_dupla2': jogo.placar_dupla2,
                    'concluido': jogo.concluido,
                } for jogo in fase_jogos
            ]
            playoff_card_style = CARD_STYLE_DICT[fase]

    data = {
        'torneio': {
            'id': torneio.id,
            'nome': torneio.nome,
            'data': torneio.data,
            'ativo': torneio.ativo,
        },
        'grupos': grupos,
        'fases_finais': fases_finais,
        'groups_finished': groups_finished,
        'can_edit': request.user.is_superuser or torneio.criado_por == request.user,
        'card_style': playoff_card_style,
    }

    return JsonResponse(data)


def export_csv(request, torneio_id: str):
    '''Export tournament data as CSV'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    jogos = Jogo.objects.filter(torneio=torneio)
    classificacao = torneio.get_groups_ranking()

    wb = Workbook()
    ws = wb.active
    ws.title = torneio.nome

    # Estilos de formatação
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    group_font = Font(bold=True, size=13)
    normal_font = Font(size=10)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    orange_fill = PatternFill(start_color='FF972F', end_color='FF972F', fill_type='solid')
    game_headers = ['', 'Dupla 1', 'Placar', '', '', 'Dupla 2']
    ranking_headers = ['#', 'Dupla', 'Vitórias', 'Saldo', 'Pontos', 'Jogos']
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center')

    # LINHA 1: Nome do torneio
    ws.merge_cells('A1:N1')
    ws['A1'] = torneio.nome
    ws['A1'].font = title_font
    ws['A1'].fill = orange_fill
    ws['A1'].alignment = center_alignment
    ws['A1'].border = thin_border
    ws['N1'].border = thin_border
    rd = ws.row_dimensions[1]
    rd.height = 45

    image_path = settings.BASE_DIR / 'setup/static/images/trophy.png'
    if os.path.exists(image_path):
        try:
            img = Image(image_path)
            img.width = 60
            img.height = 60
            ws.add_image(img, 'A1')
        except Exception as e:
            print(f"Erro ao carregar imagem: {e}")

    # LINHA 2: Data do torneio
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Data: {torneio.data.strftime('%d/%m/%Y')}"
    ws['A2'].font = header_font
    ws['A2'].fill = orange_fill
    ws['A2'].alignment = center_alignment
    ws['A2'].border = thin_border
    ws['F2'].border = thin_border

    ws.merge_cells('G2:H2')
    ws['G2'].fill = orange_fill

    ws.merge_cells('I2:N2')
    ws['I2'] = f"Jogos: {jogos.count()}"
    ws['I2'].font = header_font
    ws['I2'].fill = orange_fill
    ws['I2'].alignment = center_alignment
    ws['I2'].border = thin_border
    ws['N2'].border = thin_border

    # Processar grupos
    current_row = 5
    for i in range(1, 5):
        grupo_jogos = jogos.filter(fase=f'GRUPO {i}')
        if grupo_jogos.exists():
            # Header do grupo com merge (A:N)
            ws.merge_cells(f'A{current_row}:N{current_row}')
            ws.cell(row=current_row, column=1, value=f'GRUPO {i}')
            ws.cell(row=current_row, column=1).font = group_font
            ws.cell(row=current_row, column=1).fill = orange_fill
            ws.cell(row=current_row, column=1).alignment = center_alignment
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=14).border = thin_border
            rd = ws.row_dimensions[current_row]
            rd.height = 30
            current_row += 1

            # Headers das colunas - Jogos e Ranking lado a lado
            # Colunas A-F: Jogos | Colunas I-N: Ranking
            for col, header in enumerate(game_headers, 1):  # A, B, C, D, E, F
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            ws.merge_cells(
                start_column=3,
                start_row=current_row,
                end_column=5,
                end_row=current_row
            )

            # Headers do ranking
            for col, header in enumerate(ranking_headers, 9):  # I, J, K, L, M, N
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            current_row += 1

            # Dados dos jogos e ranking lado a lado
            games_list = list(grupo_jogos)
            ranking_list = classificacao[f'GRUPO {i}']
            max_rows = max(len(games_list), len(ranking_list))

            for row in range(max_rows):
                # Dados do jogo (colunas A-F)
                if row < len(games_list):
                    jogo = games_list[row]
                    game_data = [
                        jogo.get_concluido_display(),
                        jogo.dupla1.render_special() if jogo.dupla1 else 'A definir',
                        f'{jogo.placar_dupla1 or ""}',
                        'x',
                        f'{jogo.placar_dupla2 or ""}',
                        jogo.dupla2.render_special() if jogo.dupla2 else 'A definir',
                    ]

                    for col, data in enumerate(game_data, 1):  # A, B, C, D, E, F
                        cell = ws.cell(row=current_row, column=col, value=data)
                        cell.font = normal_font
                        cell.border = thin_border
                        cell.alignment = center_alignment

                    rd = ws.row_dimensions[current_row]
                    rd.height = 30

                # Dados do ranking (colunas I-N)
                if row < len(ranking_list):
                    dupla = ranking_list[row]
                    ranking_data = [
                        dupla['posicao'],
                        dupla['dupla'].render_special(),
                        dupla['vitorias'],
                        dupla['saldo'],
                        dupla['pontos'],
                        dupla['jogos']
                    ]

                    for col, data in enumerate(ranking_data, 9):  # I, J, K, L, M, N
                        cell = ws.cell(row=current_row, column=col, value=data)
                        cell.font = normal_font
                        cell.border = thin_border
                        cell.alignment = center_alignment

                current_row += 1

            # Linhas em branco após cada grupo
            current_row += 2

    # Processar fases de playoff
    for fase in ['OITAVAS', 'QUARTAS', 'SEMIFINAIS', 'TERCEIRO LUGAR', 'FINAL']:
        fase_jogos = jogos.filter(fase=fase)
        if fase_jogos.exists():
            # Header da fase com merge (A:N)
            ws.merge_cells(f'A{current_row}:N{current_row}')
            ws.cell(row=current_row, column=1, value=fase)
            ws.cell(row=current_row, column=1).font = group_font
            ws.cell(row=current_row, column=1).fill = orange_fill
            ws.cell(row=current_row, column=1).alignment = center_alignment
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=14).border = thin_border
            rd = ws.row_dimensions[current_row]
            rd.height = 30
            current_row += 1

            # Headers das colunas (apenas jogos, centralizados)
            playoff_headers = ['', 'Dupla 1', 'Placar', '', '', 'Dupla 2']
            start_col = 1  # Centralizar as 4 colunas (D, E, F, G)
            for col, header in enumerate(playoff_headers, start_col):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            ws.merge_cells(
                start_column=3,
                start_row=current_row,
                end_column=5,
                end_row=current_row
            )
            current_row += 1

            # Dados dos jogos
            for jogo in fase_jogos:
                game_data = [
                    jogo.get_concluido_display(),
                    jogo.dupla1.render_special() if jogo.dupla1 else 'A definir',
                    f'{jogo.placar_dupla1 or ""}',
                    'x',
                    f'{jogo.placar_dupla2 or ""}',
                    jogo.dupla2.render_special() if jogo.dupla2 else 'A definir',
                ]

                for col, data in enumerate(game_data, start_col):
                    cell = ws.cell(row=current_row, column=col, value=data)
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.alignment = center_alignment

                rd = ws.row_dimensions[current_row]
                rd.height = 30
                current_row += 1

            # Linhas em branco após cada fase
            current_row += 2

    # Ajustar largura das colunas
    column_widths = {
        'A': 5,  # Status
        'B': 15, # Dupla 1
        'C': 3,  # Placar 1,
        'D': 3,  # x
        'E': 3,  # Placar 2
        'F': 15, # Dupla 2
        'G': 3,  #
        'H': 3,  #
        'I': 3,  # Posição ranking
        'J': 15, # Dupla ranking
        'K': 8,  # Vitórias
        'L': 8,  # Pontos
        'M': 8,  # Saldo
        'N': 8,  # Jogos
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Criar response
    filename = f'{torneio.nome}_{torneio.data.strftime("%d-%m-%Y")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response