import base64, io, os, qrcode, urllib
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models import Jogo, Torneio


def distribute_classifieds(classifieds):
    '''Distribui classificados [1,2,3,4,5,6,7,8] -> [1,8,3,6,5,4,7,2]'''
    result = []
    for i in range(0, len(classifieds), 2):
        result.extend([classifieds[i], classifieds[-i-1]])
    return result


@login_required(redirect_field_name='next', login_url='/admin/login/')
def create_games(request, torneio_id: str):
    '''Cria jogos para o torneio'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    result = torneio.create_games()
    if isinstance(result, Exception):
        messages.add_message(request, messages.ERROR, result)
    else:
        messages.add_message(request, messages.INFO, 'Jogos gerados com sucesso!')
    response = redirect('admin:cup_torneio_change', torneio_id)
    response['location'] += '#jogos-tab'
    return response


def finish_tournament(request, torneio_id: str):
    '''Finaliza torneio e desabilita todas as duplas'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    torneio.finish()
    messages.add_message(request, messages.SUCCESS, f'{ torneio } finalizado!')
    response = redirect('admin:cup_torneio_change', torneio_id)
    response['location'] += '#jogos-tab'
    return response


@login_required(redirect_field_name='next', login_url='/admin/login/')
def next_stage(request, torneio_id: str):
    '''Processa grupos e vai para a próxima fase'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    jogos = torneio.jogo_set.all()
    jogos_preenchidos = jogos.filter(
        fase__in=['OITAVAS', 'QUARTAS', 'SEMIFINAIS', 'FINAL'],
        dupla1__isnull=False,
        dupla2__isnull=False
    )
    jogos_grupos_nao_finalizados = jogos.filter(fase__startswith='GRUPO').exclude(concluido='C')

    if jogos_grupos_nao_finalizados.exists():
        messages.add_message(request, messages.ERROR, 'Jogos de grupos ainda não foram finalizados.')
    elif jogos_preenchidos:
        result = torneio.next_stage()
        if isinstance(result, Exception):
            messages.add_message(request, messages.ERROR, result)
        else:
            messages.add_message(request, messages.INFO, 'Confrontos gerados com sucesso!')
    else:
        classificados = torneio.process_groups()
        classificados = distribute_classifieds(classificados)
        for i in range(0, len(classificados), 2):
            dupla1, dupla2 = classificados[i:i+2]
            if torneio.quantidade_grupos == 1:
                final = jogos.get(fase='FINAL')
                final.dupla1 = dupla1['dupla']
                final.dupla2 = dupla2['dupla']
                final.save()
            elif torneio.quantidade_grupos == 2:
                semifinal = jogos.filter(fase='SEMIFINAIS')[i//2]
                semifinal.dupla1 = dupla1['dupla']
                semifinal.dupla2 = dupla2['dupla']
                semifinal.save()
            elif torneio.quantidade_grupos == 4:
                quartas = jogos.filter(fase='QUARTAS')[i//2]
                quartas.dupla1 = dupla1['dupla']
                quartas.dupla2 = dupla2['dupla']
                quartas.save()
            elif torneio.quantidade_grupos == 8:
                oitavas = jogos.filter(fase='OITAVAS')[i//2]
                oitavas.dupla1 = dupla1['dupla']
                oitavas.dupla2 = dupla2['dupla']
                oitavas.save()
        messages.add_message(request, messages.INFO, 'Confrontos gerados com sucesso!')

    response = redirect('admin:cup_torneio_change', torneio_id)
    response['location'] += '#jogos-tab'
    return response


def see_tournament(request, torneio_id: str):
    '''Visualiza torneio'''
    CARD_STYLE_DICT = {
        'OITAVAS': 'col-sm-3',
        'QUARTAS': 'col-sm-4',
        'SEMIFINAIS': 'col-sm-6',
        'FINAL': 'col-sm-6',
    }
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    jogos = Jogo.objects.filter(torneio=torneio)
    classificacao = torneio.get_groups_ranking()
    playoff_card_style = ''

    # Separar jogos por grupos
    grupos = {}
    groups_finished = True
    for i in range(1, 5):  # Grupos 1 a 4
        grupo_jogos = jogos.filter(fase=f'GRUPO {i}')
        if grupo_jogos.exists():
            grupos[f'GRUPO {i}'] = {
                'jogos': grupo_jogos,
                'classificacao': classificacao[f'GRUPO {i}']
            }
            not_finished = grupo_jogos.exclude(concluido='C').exists()
            groups_finished = groups_finished and not not_finished

    # Separar fases finais (semifinais e final)
    fases_finais = {}
    for fase in ['FINAL', 'SEMIFINAIS', 'QUARTAS', 'OITAVAS']:
        fase_jogos = jogos.filter(fase=fase)
        if fase_jogos.exists():
            fases_finais[fase] = fase_jogos
            playoff_card_style = CARD_STYLE_DICT[fase]

    context = {
        'torneio': torneio,
        'grupos': grupos,
        'fases_finais': fases_finais,
        'classificacao': classificacao,
        'groups_finished': groups_finished,
        'playoff_card_style': playoff_card_style,
        'can_edit': request.user.is_superuser or torneio.criado_por == request.user,
    }
    return render(request, 'cup/see_cup.html', context)


@login_required(redirect_field_name='next', login_url='/admin/login/')
def qrcode_tournament(request, torneio_id: str):
    '''Cria QR Code do torneio'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    img = qrcode.make(f'{ os.getenv("APP_LINK") }/torneio-v2/{ torneio.slug }')
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    string = base64.b64encode(buf.read())
    uri =  urllib.parse.quote(string)
    context = {
        'img_b64': uri,
        'torneio': torneio,
    }
    return render(request, 'qrcode_tournament.html', context)


def get_tournament_data(request, torneio_id: str):
    '''Returns tournament data as JSON for React frontend'''
    CARD_STYLE_DICT = {
        'OITAVAS': '1/4',
        'QUARTAS': '1/3',
        'SEMIFINAIS': '1/2',
        'TERCEIRO LUGAR': '1/2',
        'FINAL': '1/2',
    }
    torneio = Torneio.objects.filter(slug=torneio_id).first()
    if not torneio:
        return JsonResponse({'error': 'Torneio não encontrado'}, status=404)
    jogos = Jogo.objects.filter(torneio=torneio)
    classificacao = torneio.get_groups_ranking()
    playoff_card_style = ''

    # Process groups
    grupos = {}
    groups_finished = True
    for i in range(1, 5):
        grupo_jogos = jogos.filter(fase=f'GRUPO {i}')
        if grupo_jogos.exists():
            grupos[f'GRUPO {i}'] = {
                'jogos': [
                    {
                        'id': jogo.id,
                        'dupla1': jogo.dupla1.render() if jogo.dupla1 else None,
                        'dupla2': jogo.dupla2.render() if jogo.dupla2 else None,
                        'pontos_dupla1': jogo.pontos_dupla1,
                        'pontos_dupla2': jogo.pontos_dupla2,
                        'concluido': jogo.concluido,
                        'obs': jogo.obs,
                    } for jogo in grupo_jogos
                ],
                'classificacao': [
                    {
                        'posicao': dupla['posicao'],
                        'dupla': dupla['dupla'].render(),
                        'vitorias': dupla['vitorias'],
                        'pontos': dupla['pontos'],
                        'saldo': dupla['saldo'],
                        'jogos': dupla['jogos'],
                    } for dupla in classificacao[f'GRUPO {i}']
                ]
            }
            groups_finished = groups_finished and not grupo_jogos.exclude(concluido='C').exists()

    # Process playoffs
    fases_finais = {}
    for fase in ['FINAL', 'TERCEIRO LUGAR', 'SEMIFINAIS', 'QUARTAS', 'OITAVAS']:
        fase_jogos = jogos.filter(fase=fase)
        if fase_jogos.exists():
            fases_finais[fase] = [
                {
                    'id': jogo.id,
                    'dupla1': jogo.dupla1.render() if jogo.dupla1 else None,
                    'dupla2': jogo.dupla2.render() if jogo.dupla2 else None,
                    'pontos_dupla1': jogo.pontos_dupla1,
                    'pontos_dupla2': jogo.pontos_dupla2,
                    'concluido': jogo.concluido,
                    'obs': jogo.obs
                } for jogo in fase_jogos
            ]
            playoff_card_style = CARD_STYLE_DICT[fase]

    data = {
        'torneio': {
            'id': torneio.id,
            'nome': torneio.nome,
            'data': torneio.data,
            'tipo': torneio.tipo,
            'ativo': torneio.ativo,
            'duplas': torneio.duplas.count(),
            'jogos': jogos.count(),
            'jogos_restantes': jogos.exclude(concluido='C').count(),
        },
        'grupos': grupos,
        'fases_finais': fases_finais,
        'groups_finished': groups_finished,
        'can_edit': torneio.ativo and (request.user.is_superuser or torneio.criado_por == request.user),
        'card_style': playoff_card_style,
    }

    return JsonResponse(data)


def export_csv(request, torneio_id: str):
    '''Export tournament data as CSV'''
    torneio = get_object_or_404(Torneio, pk=torneio_id)
    jogos = Jogo.objects.filter(torneio=torneio)
    classificacao = torneio.get_groups_ranking()

    wb = Workbook()
    ws = wb.active
    ws.title = torneio.nome

    # Estilos de formatação
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    group_font = Font(bold=True, size=13)
    normal_font = Font(size=10)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    orange_fill = PatternFill(start_color='FF972F', end_color='FF972F', fill_type='solid')
    game_headers = ['', torneio.team_title(1), 'Placar', '', '', torneio.team_title(2)]
    ranking_headers = ['#', torneio.team_title(), 'Vitórias', 'Saldo', 'Pontos', 'Jogos']
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('B2:O2')
    ws['B2'] = torneio.nome
    ws['B2'].font = title_font
    ws['B2'].fill = header_fill
    ws['B2'].alignment = center_alignment
    ws['B2'].border = thin_border
    ws['O2'].border = thin_border
    rd = ws.row_dimensions[2]
    rd.height = 45

    # Logo
    image_path = settings.BASE_DIR / 'setup/static/images/podio.png'
    if os.path.exists(image_path):
        try:
            img = Image(image_path)
            img.width = 45*4
            img.height = 45
            ws.add_image(img, 'B2')
        except Exception as e:
            print(f"Erro ao carregar imagem: {e}")

    # LINHA 3: Data do torneio
    ws.merge_cells('B3:G3')
    ws['B3'] = f"Data: {torneio.data.strftime('%d/%m/%Y')}"
    ws['B3'].font = header_font
    ws['B3'].fill = header_fill
    ws['B3'].alignment = center_alignment
    ws['B3'].border = thin_border
    ws['G3'].border = thin_border

    ws.merge_cells('H3:I3') # Espaço em branco
    ws['H3'].fill = header_fill
    ws['H3'].border = thin_border

    # Número de jogos
    ws.merge_cells('J3:O3')
    ws['J3'] = f"Jogos: {jogos.count()}"
    ws['J3'].font = header_font
    ws['J3'].fill = header_fill
    ws['J3'].alignment = center_alignment
    ws['J3'].border = thin_border
    ws['O3'].border = thin_border

    # Ajustar a linha inicial de conteúdo para a nova estrutura
    current_row = 6

    # Processar grupos
    for i in range(1, 5):
        grupo_jogos = jogos.filter(fase=f'GRUPO {i}')
        if grupo_jogos.exists():
            ws.merge_cells(f'B{current_row}:O{current_row}')
            ws.cell(row=current_row, column=2, value=f'GRUPO {i}')
            ws.cell(row=current_row, column=2).font = group_font
            ws.cell(row=current_row, column=2).fill = orange_fill
            ws.cell(row=current_row, column=2).alignment = center_alignment
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=15).border = thin_border
            rd = ws.row_dimensions[current_row]
            rd.height = 30
            current_row += 1

            # Headers das colunas - Jogos e Ranking lado a lado
            # Colunas B-G: Jogos | Colunas J-O: Ranking
            for col, header in enumerate(game_headers, 2):  # B, C, D, E, F, G
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            ws.merge_cells(
                start_column=4,
                start_row=current_row,
                end_column=6,
                end_row=current_row
            )

            # Headers do ranking
            for col, header in enumerate(ranking_headers, 10):  # J, K, L, M, N, O
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            current_row += 1

            # Dados dos jogos e ranking lado a lado
            games_list = list(grupo_jogos)
            ranking_list = classificacao[f'GRUPO {i}']
            max_rows = max(len(games_list), len(ranking_list))

            for row in range(max_rows):
                # Dados do jogo (colunas B-G)
                if row < len(games_list):
                    jogo = games_list[row]
                    game_data = [
                        jogo.get_concluido_display(),
                        jogo.dupla1.render_special() if jogo.dupla1 else 'A definir',
                        f'{jogo.pontos_dupla1 or ""}',
                        'x',
                        f'{jogo.pontos_dupla2 or ""}',
                        jogo.dupla2.render_special() if jogo.dupla2 else 'A definir',
                    ]

                    for col, data in enumerate(game_data, 2):
                        cell = ws.cell(row=current_row, column=col, value=data)
                        cell.font = normal_font
                        cell.border = thin_border
                        cell.alignment = center_alignment

                    rd = ws.row_dimensions[current_row]
                    rd.height = 30

                # Dados do ranking (colunas J-O)
                if row < len(ranking_list):
                    dupla = ranking_list[row]
                    ranking_data = [
                        dupla['posicao'],
                        dupla['dupla'].render_special(),
                        dupla['vitorias'],
                        dupla['saldo'],
                        dupla['pontos'],
                        dupla['jogos']
                    ]

                    for col, data in enumerate(ranking_data, 10):
                        cell = ws.cell(row=current_row, column=col, value=data)
                        cell.font = normal_font
                        cell.border = thin_border
                        cell.alignment = center_alignment

                current_row += 1

            # Linhas em branco após cada grupo
            current_row += 2

    # Processar fases de playoff
    for fase in ['OITAVAS', 'QUARTAS', 'SEMIFINAIS', 'TERCEIRO LUGAR', 'FINAL']:
        fase_jogos = jogos.filter(fase=fase)
        if fase_jogos.exists():
            # Header da fase com merge (B:O)
            ws.merge_cells(f'B{current_row}:O{current_row}')
            ws.cell(row=current_row, column=2, value=fase)
            ws.cell(row=current_row, column=2).font = group_font
            ws.cell(row=current_row, column=2).fill = orange_fill
            ws.cell(row=current_row, column=2).alignment = center_alignment
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=15).border = thin_border
            rd = ws.row_dimensions[current_row]
            rd.height = 30
            current_row += 1

            # Headers das colunas (apenas jogos)
            playoff_headers = ['', torneio.team_title(1), 'Placar', '', '', torneio.team_title(2)]
            start_col = 2
            for col, header in enumerate(playoff_headers, start_col):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            ws.merge_cells(
                start_column=4,
                start_row=current_row,
                end_column=6,
                end_row=current_row
            )
            current_row += 1

            # Dados dos jogos
            for jogo in fase_jogos:
                game_data = [
                    jogo.get_concluido_display(),
                    jogo.dupla1.render_special() if jogo.dupla1 else 'A definir',
                    f'{jogo.pontos_dupla1 or ""}',
                    'x',
                    f'{jogo.pontos_dupla2 or ""}',
                    jogo.dupla2.render_special() if jogo.dupla2 else 'A definir',
                ]

                for col, data in enumerate(game_data, start_col):
                    cell = ws.cell(row=current_row, column=col, value=data)
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.alignment = center_alignment

                rd = ws.row_dimensions[current_row]
                rd.height = 30
                current_row += 1

            # Linhas em branco após cada fase
            current_row += 2

    # Ajustar largura das colunas
    column_widths = {
        'A': 5,  #
        'B': 5,  # Status
        'C': 15, # Dupla 1
        'D': 3,  # Placar 1,
        'E': 3,  # x
        'F': 3,  # Placar 2
        'G': 15, # Dupla 2
        'H': 3,  #
        'I': 3,  #
        'J': 3,  # Posição ranking
        'K': 15, # Dupla ranking
        'L': 8,  # Vitórias
        'M': 8,  # Pontos
        'N': 8,  # Saldo
        'O': 8,  # Jogos
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Criar response
    filename = f'{torneio.nome}_{torneio.data.strftime("%d-%m-%Y")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response